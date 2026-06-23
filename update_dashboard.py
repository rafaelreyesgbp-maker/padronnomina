"""
update_dashboard.py
Descarga los XLSX de la carpeta "Padron de nomina" en Google Drive,
calcula el cruce IMSS-Nómina y actualiza los datos embebidos en
audit_dashboard.html. Se ejecuta por GitHub Actions (lunes-viernes).
"""

import io
import json
import re
import sys
import zipfile
import requests
import openpyxl

# ── Configuración ────────────────────────────────────────────
FOLDER_ID  = '1QoOP4jnPmp7_x9tz1Q7l2Yy2JmOx7MXc'
API_KEY    = 'AIzaSyAId7gthv7EEzmaTrfbt07FK4Kf-ii51uM'
HTML_FILE  = 'audit_dashboard.html'
BASE_URL   = 'https://www.googleapis.com/drive/v3'

# ── Regex RFC ────────────────────────────────────────────────
RFC_RE = re.compile(r'^[A-ZÑ&]{3,4}[0-9]{6}[A-Z0-9]{3}$')


def get_type(name: str) -> str | None:
    """Detecta el tipo de padrón ignorando la fecha en el nombre."""
    c = re.sub(r'\d{6,8}', '', name).lower()
    if 'imss'        in c: return 'imss'
    if re.search(r'n[oó]mina', c): return 'nomina'
    if 'cedular emp' in c or 'cedularerp' in c: return 'cedular_emp'
    if 'cedular'     in c: return 'cedular'
    if 'hospedaje'   in c: return 'hospedaje'
    if 'profesional' in c: return 'profesional'
    if re.search(r'gases?', c): return 'gases'
    if 'agua'        in c: return 'agua'
    return None


def list_files() -> list[dict]:
    """Lista los XLSX de la carpeta en Drive."""
    url = (
        f"{BASE_URL}/files"
        f"?q='{FOLDER_ID}'+in+parents+and+trashed=false"
        f"&fields=files(id,name,modifiedTime)"
        f"&key={API_KEY}&pageSize=50"
    )
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    files = r.json().get('files', [])
    return [f for f in files if get_type(f['name'])]


def download_xlsx(file_id: str) -> bytes:
    """Descarga el contenido binario de un archivo de Drive."""
    url = f"{BASE_URL}/files/{file_id}?alt=media&key={API_KEY}"
    r = requests.get(url, timeout=120)
    r.raise_for_status()
    return r.content


def extract_rfcs_xlsx(content: bytes) -> set[str]:
    """Extrae RFCs de un archivo XLSX usando openpyxl."""
    rfcs = set()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return rfcs
        # Encontrar columna RFC
        rfc_col = None
        for i, h in enumerate(header):
            if h and 'rfc' in str(h).lower():
                rfc_col = i
                break
        if rfc_col is None:
            # Buscar en primeras filas
            for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
                for i, v in enumerate(row):
                    if v and RFC_RE.match(str(v).strip().upper()):
                        rfc_col = i
                        break
                if rfc_col is not None:
                    break
        if rfc_col is None:
            rfc_col = 0

        wb2 = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws2 = wb2.active
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if rfc_col < len(row) and row[rfc_col]:
                v = str(row[rfc_col]).strip().upper()
                if RFC_RE.match(v):
                    rfcs.add(v)
    except Exception as e:
        print(f'  ⚠ Error leyendo XLSX: {e}', file=sys.stderr)
    return rfcs


def extract_rfcs_zip(content: bytes) -> set[str]:
    """Extrae RFCs desde sharedStrings.xml dentro del XLSX (ZIP)."""
    rfcs = set()
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as z:
            shared = next((n for n in z.namelist() if 'sharedstrings' in n.lower()), None)
            if shared:
                text = z.read(shared).decode('utf-8', errors='ignore')
                for m in RFC_RE.finditer(text):   # no funciona así — ver abajo
                    rfcs.add(m.group())
    except Exception:
        pass
    return rfcs


def extract_rfcs(content: bytes) -> set[str]:
    """Intenta con openpyxl; si falla, cae en regex sobre sharedStrings."""
    rfcs = extract_rfcs_xlsx(content)
    if not rfcs:
        # Fallback: regex en sharedStrings.xml
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                shared = next((n for n in z.namelist() if 'sharedstrings' in n.lower()), None)
                if shared:
                    text = z.read(shared).decode('utf-8', errors='ignore')
                    pattern = re.compile(r'[A-ZÑ&]{3,4}[0-9]{6}[A-Z0-9]{3}')
                    rfcs = set(pattern.findall(text))
        except Exception as e:
            print(f'  ⚠ Fallback también falló: {e}', file=sys.stderr)
    return rfcs


def compute_cross(rfc_sets: dict) -> tuple[list[dict], dict]:
    """Calcula IMSS sin Nómina y cruza con demás padrones."""
    imss   = rfc_sets.get('imss',        set())
    nomina = rfc_sets.get('nomina',       set())
    ced    = rfc_sets.get('cedular',      set())
    hos    = rfc_sets.get('hospedaje',    set())
    pro    = rfc_sets.get('profesional',  set())
    gas    = rfc_sets.get('gases',        set())
    agua   = rfc_sets.get('agua',         set())

    sin_nomina = [r for r in imss if r not in nomina]
    n = len(sin_nomina)

    rows = [{
        'rfc':         rfc,
        'nombre':      '',
        'direccion':   '',
        'trabajadores': None,
        'cedular':     rfc in ced,
        'hospedaje':   rfc in hos,
        'profesional': rfc in pro,
        'gases':       rfc in gas,
        'agua':        rfc in agua,
    } for rfc in sin_nomina]

    def pct(s):
        return round(len(s & set(sin_nomina)) / n * 100, 2) if n else 0.0

    stats = {
        'total_imss':      len(imss),
        'total_nomina':    len(nomina),
        'sin_nomina':      n,
        'pct_cedular':     pct(ced),
        'pct_hospedaje':   pct(hos),
        'pct_profesional': pct(pro),
        'pct_gases':       pct(gas),
        'pct_agua':        pct(agua),
        'total_gases':     len(gas & set(sin_nomina)),
        'total_agua':      len(agua & set(sin_nomina)),
    }
    return rows, stats


def update_html(rows: list[dict], stats: dict):
    """Reemplaza EMBEDDED_ROWS y EMBEDDED_STATS en el HTML."""
    with open(HTML_FILE, encoding='utf-8') as f:
        content = f.read()

    rows_js  = json.dumps(rows,  ensure_ascii=False, separators=(',', ':'))
    stats_js = json.dumps(stats, ensure_ascii=False, separators=(',', ':'))

    # Reemplazar línea por línea para evitar problemas con regex y JSON anidado
    lines = content.split('\n')
    new_lines = []
    for line in lines:
        if line.startswith('const EMBEDDED_ROWS='):
            new_lines.append(f'const EMBEDDED_ROWS={rows_js};')
        elif line.startswith('const EMBEDDED_STATS='):
            new_lines.append(f'const EMBEDDED_STATS={stats_js};')
        else:
            new_lines.append(line)

    with open(HTML_FILE, 'w', encoding='utf-8') as f:
        f.write('\n'.join(new_lines))

    print(f'✓ HTML actualizado: {len(rows):,} registros sin Nómina')


def main():
    print('── Auditoría Fiscal: actualización automática ──')

    print('▸ Listando archivos en Drive…')
    files = list_files()
    if not files:
        print('✗ No se encontraron archivos reconocibles. ¿La carpeta es pública?')
        sys.exit(1)

    for f in files:
        print(f"  • {f['name']} → {get_type(f['name'])}")

    rfc_sets = {}
    for f in files:
        ftype = get_type(f['name'])
        print(f"▸ Descargando {f['name']}…")
        try:
            content = download_xlsx(f['id'])
            rfcs    = extract_rfcs(content)
            rfc_sets[ftype] = rfcs
            print(f"  {len(rfcs):,} RFCs extraídos")
        except Exception as e:
            print(f'  ⚠ Error: {e}', file=sys.stderr)
            rfc_sets[ftype] = set()

    print('▸ Calculando cruce IMSS–Nómina…')
    rows, stats = compute_cross(rfc_sets)
    print(f"  IMSS: {stats['total_imss']:,} | Nómina: {stats['total_nomina']:,} | Sin Nómina: {stats['sin_nomina']:,}")

    print(f'▸ Actualizando {HTML_FILE}…')
    update_html(rows, stats)
    print('✓ Proceso completado.')


if __name__ == '__main__':
    main()
